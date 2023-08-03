import openpyxl 
import xlsxwriter
import pandas

# path = "C:\\Users\\ASUS\\OneDrive\\Documents\\marks.xlsx"
# wb_obj = openpyxl.load_workbook(path) 
# wb_result = openpyxl.Workbook()

path = "C:\\Users\\ASUS\\OneDrive\\Documents\\marks.xlsx"
wb_obj = openpyxl.load_workbook(path) 
wb_result = openpyxl.Workbook()


sheet_obj = wb_obj.active 

m_row = sheet_obj.max_row 
n=[]
rollno=[]
s=[]
m=[]
h=[]
e=[]
ss=[]
num=[]
# Loop will print all values 
# of first column  
for i in range(2, m_row + 1): 
    d1={}
    name = sheet_obj.cell(row = i, column = 1).value
    n.append(name)
    roll_no = sheet_obj.cell(row = i, column = 2).value
    rollno.append(roll_no)
    sci = sheet_obj.cell(row = i, column = 3).value
    s.append(sci)
    maths = sheet_obj.cell(row = i, column = 4).value
    m.append(maths)
    hindi = sheet_obj.cell(row = i, column = 5).value
    h.append(hindi)
    eng = sheet_obj.cell(row = i, column = 6).value
    e.append(eng)
    social = sheet_obj.cell(row = i, column = 7).value
    ss.append(social)
    d1={"Name":name,"Roll_no":roll_no,"Science":sci,"Maths":maths,"Hindi":hindi,"English":eng,"Social Science":social}
    num.append(d1)
    print(num)

def show_all_records():
    df = pandas.DataFrame(num)
    writer = pandas.ExcelWriter( 'all_records.xlsx',  
                   engine ='xlsxwriter') 
  
    df.to_excel(writer, sheet_name ='Sheet1') 
  

    writer._save() 

    
def searchbyname():
    search=str(input("Enter the name you want to search:"))
    flag = False
    for j in range(0,len(num)):
        if search in num[j].values():
            nn=[]
            nn = num[j].values()
            text = "Hurrah we found your searched value!!"
            df = pandas.DataFrame(nn)
            writer = pandas.ExcelWriter('searchbyName.xlsx',  
                   engine ='xlsxwriter') 
  
            df.to_excel(writer, startrow=1, startcol=0) 
            
            worksheet = writer.sheets['Sheet1']
            worksheet.write(0, 0, text)
              

            writer._save()  
            
            
            flag = False
            break
        else:
            flag = True
    if(flag == True):
        print("Record is not found")   
        
def searchingbyrollnumber():
    value2 = int(input("enter the roll number:"))
    flag1 = False
    for k in range(0,len(num)):   
        if value2 in num[k].values():
            rr=[] 
            rr = num[k].values()
            text1 = "Hurrah we found your searched value!!"
            df = pandas.DataFrame(rr)
            writer = pandas.ExcelWriter('searchbyRollNo.xlsx',  
                   engine ='xlsxwriter') 
  
            df.to_excel(writer, startrow=1, startcol=0) 
            
            worksheet = writer.sheets['Sheet1']
            worksheet.write(0, 0, text1)
            writer._save()  
            flag1 = False
            break
        else:
            flag1 = True
    if(flag1 == True):
        print("Record is not found")  

def result():
    search=str(input("Enter the name you want to search:"))
    flag = False
    k = {'Hurrah we found student': ' '}
    no =[]
    no.append(k)
    for v in range(0,len(num)):
        if search in num[v].values():
            sum = []
            total = 5*100
            sum = num[v]["Science"] + num[v]["Maths"] + num[v]["Hindi"] + num[v]["English"] + num[v]["Social Science"]
            text1= "Student got "
            text2 = "Marks  out of 500 "
            df = pandas.DataFrame(no)
            writer = pandas.ExcelWriter('result.xlsx',  
                   engine ='xlsxwriter') 
  
            df.to_excel(writer, startrow=1, startcol=0) 
            
            worksheet = writer.sheets['Sheet1']
            worksheet.write(2, 1, text1)
            worksheet.write(2, 2, sum)
            worksheet.write(2, 3, text2)
            print("Student got" , int(sum) , "Marks  out of " , int(total))
            
            per = sum/total*100
            print(per)
            if(per>60):
                txt= "Student got First Position"
                worksheet.write(3, 1, txt)
            elif(per>50):
                txt= "Student got Second Position"
                worksheet.write(3, 1, txt)
               
            elif(per>33):
                txt= "Student got Third Position"
                worksheet.write(3, 1, txt)
                
            else:
                txt= "WELLDONE DEAR STUDENT YOU PASSES WITH DISTINCTION"
                worksheet.write(3, 1, txt)

            writer._save()  
            flag = False
            break
        else:
            flag = True    
    if(flag == True):
        
        print("SORRY! No Student Record  found")
        
def topperm():
     k = {' ': ' '}
     no =[]
     no.append(k)
     max(m)
     print(max(m))
     search=max(m)
     for  g in range(0,len(num)):   
         if search  in num[g].values():
             txt= "Topper of maths is"
             name_t = num[g]["Name"]
             txt1= "roll no."
             rollno_t = num[g]["Roll_no"]
             txt2= "with numbers"
             marks_t = num[g]["Maths"]
             df = pandas.DataFrame(no)
             writer = pandas.ExcelWriter('maths_topper.xlsx',  
                   engine ='xlsxwriter') 
  
             df.to_excel(writer, startrow=1, startcol=0) 
            
             worksheet = writer.sheets['Sheet1']
             worksheet.write(2, 1, txt)
             worksheet.write(2, 2, name_t)
             worksheet.write(3, 1, txt1)
             worksheet.write(3, 2, rollno_t)
             worksheet.write(4, 1, txt2)
             worksheet.write(4, 2, marks_t)
             writer._save() 
             
             
             print("topper of maths is",num[g]["Name"])
             print("Roll number",num[g]["Roll_no"])
             print("with numbers",num[g]["Maths"])
    
def toppere():    
    k = {' ': ' '}
    no =[]
    no.append(k) 
    max1 = max(e)
    search = str(max1)

    for g in range(0,len(num)):

        if search in str(num[g].get("English")):
             txt= "Topper of Eng is"
             name_t = num[g]["Name"]
             txt1= "roll no."
             rollno_t = num[g]["Roll_no"]
             txt2= "with numbers"
             marks_t = num[g]["English"]
             df = pandas.DataFrame(no)
             writer = pandas.ExcelWriter('english_topper.xlsx',  
                   engine ='xlsxwriter') 
  
             df.to_excel(writer, startrow=1, startcol=0) 
            
             worksheet = writer.sheets['Sheet1']
             worksheet.write(2, 1, txt)
             worksheet.write(2, 2, name_t)
             worksheet.write(3, 1, txt1)
             worksheet.write(3, 2, rollno_t)
             worksheet.write(4, 1, txt2)
             worksheet.write(4, 2, marks_t)
             writer._save() 
            # print("topper of English is",num[g]["Name"])
            # print("Roll number",num[g]["Roll_no"])
            # print("with numbers",num[g]["English"])
        
def toppers():    
    k = {' ': ' '}
    no =[]
    no.append(k) 
    max1 = max(s)
    search = str(max1)

    for g in range(0,len(num)):

        if search in str(num[g].get("Science")):
             txt= "Topper of Science is"
             name_t = num[g]["Name"]
             txt1= "roll no."
             rollno_t = num[g]["Roll_no"]
             txt2= "with numbers"
             marks_t = num[g]["Science"]
             df = pandas.DataFrame(no)
             writer = pandas.ExcelWriter('sci_topper.xlsx',  
                   engine ='xlsxwriter') 
  
             df.to_excel(writer, startrow=1, startcol=0) 
            
             worksheet = writer.sheets['Sheet1']
             worksheet.write(2, 1, txt)
             worksheet.write(2, 2, name_t)
             worksheet.write(3, 1, txt1)
             worksheet.write(3, 2, rollno_t)
             worksheet.write(4, 1, txt2)
             worksheet.write(4, 2, marks_t)
             writer._save() 
            # print("topper of Science is",num[g]["Name"])
            # print("Roll number",num[g]["Roll_no"])
            # print("with numbers",num[g]["Science"])
        
    
def topperss():  
    k = {' ': ' '}
    no =[]
    no.append(k)           
    max1 = max(ss)

    search = str(max1)

    for g in range(0,len(num)):

        if search in str(num[g].get("Social Science")):
             txt= "Topper of Social Science is"
             name_t = num[g]["Name"]
             txt1= "roll no."
             rollno_t = num[g]["Roll_no"]
             txt2= "with numbers"
             marks_t = num[g]["Social Science"]
             df = pandas.DataFrame(no)
             writer = pandas.ExcelWriter('socialSci_topper.xlsx',  
                   engine ='xlsxwriter') 
  
             df.to_excel(writer, startrow=1, startcol=0) 
            
             worksheet = writer.sheets['Sheet1']
             worksheet.write(2, 1, txt)
             worksheet.write(2, 2, name_t)
             worksheet.write(3, 1, txt1)
             worksheet.write(3, 2, rollno_t)
             worksheet.write(4, 1, txt2)
             worksheet.write(4, 2, marks_t)
             writer._save() 
            # print("topper of Social Science is",num[g]["Name"])
            # print("Roll number",num[g]["Roll_no"])
            # print("with numbers",num[g]["Social Science"])
        
def topperh():   
    k = {' ': ' '}
    no =[]
    no.append(k) 
    max1 = max(h)
    search = str(max1)

    for g in range(0,len(num)):

        if search in str(num[g].get("Hindi")):
             txt= "Topper of Hindi is"
             name_t = num[g]["Name"]
             txt1= "roll no."
             rollno_t = num[g]["Roll_no"]
             txt2= "with numbers"
             marks_t = num[g]["Hindi"]
             df = pandas.DataFrame(no)
             writer = pandas.ExcelWriter('hindi_topper.xlsx',  
                   engine ='xlsxwriter') 
  
             df.to_excel(writer, startrow=1, startcol=0) 
            
             worksheet = writer.sheets['Sheet1']
             worksheet.write(2, 1, txt)
             worksheet.write(2, 2, name_t)
             worksheet.write(3, 1, txt1)
             worksheet.write(3, 2, rollno_t)
             worksheet.write(4, 1, txt2)
             worksheet.write(4, 2, marks_t)
             writer._save() 
            
while(True):
    print("enter 0 to exit \n Hit 1 to see all records \n Hit 2 to search by name \n Hit 3 to search by roll number \n Hit 4 to see result \n Hit 5 to see topper of English\n Hit 6 to see topper of Maths\n Hit 7 to see topper of Hindi\n Hit 8 to see topper of Social Science\n Hit 9 to see topper of science")
    choice = int(input("enter the number:"))
    if(choice == 0):
        break       
    elif(choice == 1):
        show_all_records()
    elif(choice==2):
        searchbyname()
    elif(choice==3):
        searchingbyrollnumber()
    elif(choice==4):
        result()
    elif(choice==5):
        toppere()
    elif(choice==6):
        topperm()
    elif(choice==7):
        topperh()
    elif(choice==8):
        topperss()
    elif(choice==9):
        toppers()
    else:
        print("Thanks for visting")

